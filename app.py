import streamlit as st
import pandas as pd
from io import BytesIO
from unidecode import unidecode
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="📊 Análise de Vendas e Consumo de Estoque", layout="centered")
st.title("📊 Sistema Integrado de Vendas e Consumo")

# ========================== AGENTE DE VENDAS ==========================
st.header("🍽️ Análise de Maiores Vendas")
file_vendas = st.file_uploader("Faça upload da planilha de VENDAS", type=["xlsx"], key="vendas")

def normalizar(texto):
    return str(texto).lower().strip()

if file_vendas:
    df_vendas = pd.read_excel(file_vendas, skiprows=3)
    df_vendas["Itens e Opções"] = df_vendas["Itens e Opções"].astype(str).apply(normalizar)

    mult = {
        "- 2 pequenos": 2, "- 3 pequenos": 3, "- 4 pequenos": 4,
        "- 2 grandes": 2, "- 3 grandes": 3, "- 4 grandes": 4
    }
    for k, m in mult.items():
        df_vendas.loc[df_vendas["Itens e Opções"].str.contains(k), "Quantidade"] *= m

    pequeno = df_vendas["Itens e Opções"].str.contains("pequeno") & ~df_vendas["Itens e Opções"].str.contains("combo")
    grande = df_vendas["Itens e Opções"].str.contains("grande") & ~df_vendas["Itens e Opções"].str.contains("combo")
    total_p = int(df_vendas.loc[pequeno, "Quantidade"].sum())
    total_g = int(df_vendas.loc[grande, "Quantidade"].sum())
    total_geral = total_p + total_g

    st.subheader("Resumo de Pequenos e Grandes")
    st.write(f"Pequeno: {total_p}")
    st.write(f"Grande: {total_g}")
    st.write(f"Total: {total_geral}")

    pratos = {
        "Boi": lambda x: "boi" in x and "combo" not in x,
        "Parmegiana": lambda x: "parmegiana" in x and "combo" not in x,
        "Strogonoff": lambda x: "strogonoff" in x and "combo" not in x,
        "Feijoada": lambda x: "feijoada" in x and "2 feijoadas" not in x,
        "Tropeiro": lambda x: "tropeiro" in x and "tropeguete" not in x,
        "Tropeguete": lambda x: "tropeguete" in x,
        "Espaguete": lambda x: "espaguete" in x and "tropeguete" not in x,
        "Porco": lambda x: "porco" in x and "combo" not in x,
        "Frango": lambda x: "frango" in x and "parmegiana" not in x and "2 frangos + fritas" not in x
    }

    combos = {
        "Combo Todo Dia": lambda x: "combo todo dia" in x,
        "2 Pratos - À Sua Escolha": lambda x: "2 pratos" in x and "escolha" in x,
        "Combo Supremo": lambda x: "combo supremo" in x,
        "2 Feijoadas": lambda x: "2 feijoadas" in x,
        "2 Frangos + Fritas": lambda x: "2 frangos" in x and "fritas" in x
    }

    refrigerantes = {
        "Coca-Cola Original 350 ml": [["coca", "original", "350"]],
        "Coca-Cola Zero e Sem Açúcar 350 ml": [["coca", "zero", "350"], ["coca", "sem acucar", "350"]],
        "Coca-Cola Original 600 ml": [["coca", "original", "600"]],
        "Coca-Cola Zero 600 ml": [["coca", "zero", "600"], ["coca", "sem acucar", "600"]],
        "Coca-Cola 2 Litros": [["coca", "2l"], ["coca", "2 l"], ["coca", "2litro"]],
        "Guaraná Antarctica 350 ml": [["guarana", "350"]],
        "Guaraná Antarctica 1 Litro": [["guarana", "antarctica", "1l"], ["guarana", "antarctica", "1 l"], ["guarana", "antarctica", "1litro"]],
        "Guaraná Antarctica 2 Litros": [["guarana", "2l"], ["guarana", "2 l"], ["guarana", "2litro"]],
        "Suco": [["suco"]],
        "Refrigerante Mate Couro 1 Litro": [["mate couro", "1l"], ["guarana mate", "1l"], ["mate couro", "1 l"], ["guarana mate", "1 l"], ["mate couro", "1litro"], ["guarana mate", "1litro"]]
    }

    def contem_tags(texto, listas):
        return any(all(tag in texto for tag in tags) for tags in listas)

    resumo = []
    for nome, cond in pratos.items():
        f = df_vendas["Itens e Opções"].apply(cond)
        qtd = int(df_vendas.loc[f, "Quantidade"].sum())
        val = df_vendas.loc[f, "Valor Total"].sum()
        if qtd > 0:
            resumo.append({"Categoria": nome, "Quantidade": qtd, "Valor Total": val})

    for nome, cond in combos.items():
        f = df_vendas["Itens e Opções"].apply(cond)
        qtd = int(df_vendas.loc[f, "Quantidade"].sum())
        val = df_vendas.loc[f, "Valor Total"].sum()
        if qtd > 0:
            resumo.append({"Categoria": nome, "Quantidade": qtd, "Valor Total": val})

    for nome, tags in refrigerantes.items():
        f = df_vendas["Itens e Opções"].apply(lambda x: contem_tags(x, tags))
        qtd = int(df_vendas.loc[f, "Quantidade"].sum())
        val = df_vendas.loc[f, "Valor Total"].sum()
        if qtd > 0:
            resumo.append({"Categoria": nome, "Quantidade": qtd, "Valor Total": val})

    resumo_df = pd.DataFrame(resumo)
    resumo_df = resumo_df.sort_values(by="Valor Total", ascending=False)
    top5 = resumo_df.nlargest(5, "Valor Total")
    def azul_escuro(val):
        return 'color: #003366' if val in top5["Valor Total"].values else 'color: black'

    st.subheader("📋 Resumo Final Agrupado")
    st.dataframe(resumo_df.style.applymap(azul_escuro, subset=["Valor Total"]), use_container_width=True)

    output = BytesIO()
    resumo_df.to_excel(output, index=False, engine='openpyxl')
    st.download_button(
        label="📅 Baixar Análise de Vendas (.xlsx)",
        data=output.getvalue(),
        file_name="analise_maiores_vendas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ========================== AGENTE DE CONSUMO ==========================
st.header("📦 Analisador de Consumo de Estoque")

HIST_DIR = "historico_relatorios"
os.makedirs(HIST_DIR, exist_ok=True)

file = st.file_uploader("📤 Envie sua planilha de estoque .xlsx", type=["xlsx"], key="estoque")

def normalizar_nome(nome):
    if pd.isna(nome): return ""
    return unidecode(str(nome)).strip().lower()

def extrair_valor(valor):
    if pd.isna(valor): return 0.0
    valor_str = str(valor).replace("R$", "").replace(" ", "").strip()
    valor_str = re.sub(r"[^0-9,\.]+", "", valor_str)
    if "," in valor_str and "." in valor_str:
        if valor_str.rfind(",") > valor_str.rfind("."):
            valor_str = valor_str.replace(".", "").replace(",", ".")
        else:
            valor_str = valor_str.replace(",", "")
    elif "," in valor_str:
        valor_str = valor_str.replace(",", ".")
    try:
        return float(valor_str)
    except:
        return 0.0

def detectar_unidade(item):
    item = item.lower()
    if any(x in item for x in ["kg", "quilo", "grama", "g"]): return "KG"
    elif any(x in item for x in ["litro", "ml"]): return "L"
    elif any(x in item for x in ["caixa", "pct", "pacote"]): return "PCT"
    else: return "UN"

def extrair_bloco_horizontal(df, col_inicio, col_fim):
    dados = []
    for i in range(1, len(df)):
        linha = df.iloc[i, col_inicio:col_fim].tolist()
        if len(linha) >= 4 and pd.notna(linha[0]):
            item = normalizar_nome(linha[0])
            try:
                qtd = extrair_valor(linha[1])
                val = extrair_valor(linha[3])
                if item:
                    dados.append({'Item': item, 'Quantidade': qtd, 'Valor total': val})
            except:
                continue
    return pd.DataFrame(dados)

def exportar_excel_formatado(df, path):
    df.to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    col_idx = [cell.column for cell in ws[1] if cell.value == "Valor consumido"]
    if col_idx:
        col_letter = get_column_letter(col_idx[0])
        top5_vals = df.nlargest(5, "Valor consumido")["Valor consumido"].tolist()
        for row in range(2, len(df)+2):
            cell = ws[f"{col_letter}{row}"]
            if cell.value in top5_vals:
                cell.font = Font(color="FF0000")
    wb.save(path)

def analisar_consumo_estoque(file):
    df = pd.read_excel(file, sheet_name=0, header=None)
    colunas = df.iloc[0].astype(str).apply(lambda x: unidecode(str(x)).strip().lower())

    col_estoque_ini_idx = colunas[colunas.str.contains("estoque.*inicial")]
    col_compras_idx = colunas[colunas.str.contains("compras")]
    col_estoque_fim_idx = colunas[colunas.str.contains("estoque.*final")]

    if col_estoque_ini_idx.empty or col_compras_idx.empty or col_estoque_fim_idx.empty:
        st.error("❌ Não foi possível localizar os blocos de dados. Verifique os títulos da planilha.")
        return None

    col_estoque_ini = col_estoque_ini_idx.index[0]
    col_compras = col_compras_idx.index[0]
    col_estoque_fim = col_estoque_fim_idx.index[0]

    estoque_inicial = extrair_bloco_horizontal(df, col_estoque_ini, col_compras)
    compras = extrair_bloco_horizontal(df, col_compras, col_estoque_fim)
    estoque_final = extrair_bloco_horizontal(df, col_estoque_fim, col_estoque_fim + 4)

    def agrupar(df):
        return df.groupby('Item', as_index=False).agg({'Quantidade': 'sum', 'Valor total': 'sum'})

    estoque_inicial = agrupar(estoque_inicial)
    compras = agrupar(compras)
    estoque_final = agrupar(estoque_final)

    ini_qtd = dict(zip(estoque_inicial['Item'], estoque_inicial['Quantidade']))
    ini_val = dict(zip(estoque_inicial['Item'], estoque_inicial['Valor total']))
    comp_qtd = dict(zip(compras['Item'], compras['Quantidade']))
    comp_val = dict(zip(compras['Item'], compras['Valor total']))
    fin_qtd = dict(zip(estoque_final['Item'], estoque_final['Quantidade']))
    fin_val = dict(zip(estoque_final['Item'], estoque_final['Valor total']))

    itens_ordem = list(set(list(ini_qtd.keys()) + list(comp_qtd.keys()) + list(fin_qtd.keys())))
    resultado = []

    for item in itens_ordem:
        if item.strip() == "" or item.lower() == "item": continue
        qtd_ini = ini_qtd.get(item, 0)
        qtd_comp = comp_qtd.get(item, 0)
        qtd_fin = fin_qtd.get(item, 0)
        val_ini = ini_val.get(item, 0)
        val_comp = comp_val.get(item, 0)
        val_fin = fin_val.get(item, 0)
        qtd_consumida = qtd_ini + qtd_comp - qtd_fin
        val_consumido = val_ini + val_comp - val_fin
        if qtd_consumida == 0 and val_consumido == 0: continue
        unidade = detectar_unidade(item)
        resultado.append({
            'Item': item,
            'Quantidade consumida': round(qtd_consumida, 2),
            'Valor consumido': round(val_consumido, 2),
            })

    resultado_df = pd.DataFrame(resultado)
    resultado_df = resultado_df.sort_values(by="Valor consumido", ascending=False)
    return resultado_df

if file:
    resultado_df = analisar_consumo_estoque(file)
    if resultado_df is not None:
        st.success("✅ Análise concluída com sucesso!")
        top5 = resultado_df.nlargest(5, "Valor consumido")
        def destaque(val):
            return 'color: red' if val in top5["Valor consumido"].values else 'color: black'
        st.dataframe(resultado_df.style.applymap(destaque, subset=["Valor consumido"]))
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        historico_path = os.path.join(HIST_DIR, f"consumo_{now}.xlsx")
        exportar_excel_formatado(resultado_df, historico_path)
        with open(historico_path, "rb") as f:
            st.download_button(
                label="⬇️ Baixar relatório Excel",
                data=f,
                file_name="relatorio_consumo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
